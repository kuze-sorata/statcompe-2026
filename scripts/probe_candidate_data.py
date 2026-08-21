"""有力テーマ候補の公式外部データを取得し、配布形式を検証する。"""

from __future__ import annotations

import csv
import hashlib
import io
import zipfile
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import requests

from statcompe_2026.paths import PROJECT_ROOT, RAW_DATA_DIR, REFERENCES_DIR

SOURCES = {
    "food-access-2020": (
        "https://www.maff.go.jp/primaff/seika/fsc/faccess/attach/excel/2020_table05.xlsx"
    ),
    "local-government-dx-2024": (
        "https://www.digital.go.jp/assets/contents/node/basic_page/field_ref_resources/"
        "51a5a201-e0dd-493f-9c21-0692402d93e6/85162d87/"
        "20240712_resources_govdashboard_local_governmentdx_table_01.zip"
    ),
    "heatstroke-2025": (
        "https://www.fdma.go.jp/disaster/heatstroke/items/heatstroke003_data_r7.xlsx"
    ),
}


@dataclass(frozen=True)
class ProbeRow:
    source_id: str
    url: str
    local_path: str
    retrieved_at: str
    bytes: int
    sha256: str
    structure: str


def inspect_xlsx(content: bytes) -> str:
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    return "; ".join(
        f"{sheet.title}:{sheet.max_row}x{sheet.max_column}" for sheet in workbook.worksheets
    )


def inspect_zip(content: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        return "; ".join(archive.namelist())


def main() -> None:
    session = requests.Session()
    session.headers["User-Agent"] = "statcompe-2026 research project"
    destination_dir = RAW_DATA_DIR.parent / "external" / "candidate-probes"
    destination_dir.mkdir(parents=True, exist_ok=True)
    retrieved_at = datetime.now(UTC).isoformat()
    rows: list[ProbeRow] = []
    for source_id, url in SOURCES.items():
        response = session.get(url, timeout=90)
        response.raise_for_status()
        content = response.content
        suffix = Path(url).suffix
        destination = destination_dir / f"{source_id}{suffix}"
        destination.write_bytes(content)
        structure = inspect_xlsx(content) if suffix == ".xlsx" else inspect_zip(content)
        rows.append(
            ProbeRow(
                source_id=source_id,
                url=url,
                local_path=destination.relative_to(PROJECT_ROOT).as_posix(),
                retrieved_at=retrieved_at,
                bytes=len(content),
                sha256=hashlib.sha256(content).hexdigest(),
                structure=structure,
            )
        )

    manifest = REFERENCES_DIR / "manifests" / "candidate-probes.csv"
    with manifest.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(asdict(rows[0])))
        writer.writeheader()
        writer.writerows(asdict(row) for row in rows)
    for row in rows:
        print(f"{row.source_id}: {row.structure}")
    print(manifest.relative_to(PROJECT_ROOT))


if __name__ == "__main__":
    main()
